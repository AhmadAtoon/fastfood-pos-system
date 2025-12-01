from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

class ExcelExporter:
    def __init__(self, filename: str):
        self.filename = filename

    def export(self, data: List[Dict[str, Any]], sheet_name: str = "Report") -> str:
        """
        داده‌ها را به فایل Excel خروجی می‌گیرد.
        data: لیست دیکشنری‌ها با کلیدهای یکسان
        sheet_name: نام شیت
        """
        if not data:
            raise ValueError("No data to export")

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # نوشتن هدرها
        headers = list(data[0].keys())
        ws.append(headers)

        # استایل هدرها
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # نوشتن داده‌ها
        for row in data:
            ws.append(list(row.values()))

        wb.save(self.filename)
        return self.filename

    def export_with_styles(self, data: List[Dict[str, Any]], sheet_name: str = "StyledReport") -> str:
        """
        خروجی Excel با استایل بیشتر (فونت، ترازبندی)
        """
        if not data:
            raise ValueError("No data to export")

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        headers = list(data[0].keys())
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True, color="0000FF")
            cell.alignment = Alignment(horizontal="center")

        for row in data:
            ws.append(list(row.values()))

        # ترازبندی داده‌ها
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="left")

        wb.save(self.filename)
        return self.filename
